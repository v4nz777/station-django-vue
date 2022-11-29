from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.dimensions import ColumnDimension
from datetime import datetime
from os import path

from .models import User, Department




def create_new_workbook():
    today = datetime.now()
    month = today.strftime("%B")
    workbook = Workbook()
    

    workbook.save(f"dtr/output/{month}_{today.year}.xlsx")
    print('WORKBOOK CREATED')



def create_dtr(worksheet, worker):
    
    today = datetime.now()


    worksheet.merge_cells('B2:H2')
    worksheet.merge_cells('B3:H3')
    worksheet.merge_cells('B4:H4')
    worksheet.merge_cells('B6:H6')
    worksheet.merge_cells('B8:H8')
    worksheet.merge_cells('B9:H9')
    worksheet.merge_cells('B11:D11')
    worksheet.merge_cells('B12:D13')


    worksheet.merge_cells('F11:H11')
    worksheet.merge_cells('E12:F12')
    worksheet.merge_cells('E13:F13')
    worksheet.merge_cells('G12:H12')
    worksheet.merge_cells('G13:H13')
    worksheet.merge_cells('C15:D15')
    worksheet.merge_cells('E15:F15')



    worksheet.merge_cells('B58:H58')
    worksheet.merge_cells('B59:H59')
    worksheet.merge_cells('B15:B16')
    worksheet.merge_cells('G15:H15')
    worksheet.merge_cells('B48:F48')
    worksheet.merge_cells('B50:H52')
    worksheet.merge_cells('B56:H56')
    


    #HEADERS, TITLES, SUBTITLES
    worksheet['B6'].value = 'DAILY TIME RECORD'
    worksheet['B9'].value = '(name)'
    worksheet['B2'].value = 'DXIQ 106.3 LOVE RADIO MALAYBALAY'
    worksheet['B3'].value = '2nd Floor Modestar Bldg., P-2 Lumbo'
    worksheet['B11'].value = 'For the Month of:'
    worksheet['B12'].value = 'Official hours for arrival and departure'
    worksheet['E12'].value = 'Regular days'
    worksheet['E13'].value = 'DAY OFFs'
    worksheet['B15'].value = 'DAY'
    worksheet['C15'].value = 'A.M.'
    worksheet['E15'].value = 'P.M.'
    worksheet['G15'].value = 'OVERTIME'
    worksheet['C16'].value = 'IN'
    worksheet['D16'].value = 'OUT'
    worksheet['E16'].value = 'IN'
    worksheet['F16'].value = 'OUT'
    worksheet['G16'].value = 'HOURS'
    worksheet['H16'].value = 'MINUTES'
    worksheet['B50'].value = 'I certify on my honor that the above is true and correct report of the hours of work performed, recorded of which was made daily at the time of arrival and departure from office.'
    worksheet['B56'].value = 'VERIFIED as to the prescribed office hours.'
    worksheet['B48'].value = 'TOTAL'
    worksheet['B59'].value = 'In - Charge'
    
    

    #NAME OF WORKER
    u = User.objects.get(username=worker)
    worksheet['B8'].value = f'{u.first_name.title()} {u.last_name.title()}'

    #NAME OF DTR IN CHARGE
    #TODO
    for dept in Department.objects.all():
        if dept.title.lower() == 'human resource' and dept.head != None:
            worksheet['B58'].value = f'{dept.head.first_name.title()} {dept.head.last_name.title()}'
        elif dept.title == '':
            worksheet['B58'].value = 'DTR IN-CHARGE NOT SET'
        else:
            pass
            manager = User.objects.get(position=1)
            worksheet['B58'].value = f'{manager.first_name.title()} {manager.last_name.title()}'
    

    #MONTH AND YEAR
    worksheet['F11'].value = today.strftime('%B %Y')
    
    
    #CREATE DATES
    days_num = 1
    for n in range(1,32):
        days_num = n
        worksheet[f'B{16 + days_num}'].value = days_num


    round_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    under_border = Border(left=Side(style='none'), 
                         right=Side(style='none'), 
                         top=Side(style='none'), 
                         bottom=Side(style='thin'))


    #BORDERS FOR FILLABLES
    for _row in range(15,49):
        for _column in range(2,9):
            worksheet.cell(row=_row, column=_column).border = round_border

    #UNDERLINES
    for _row in range(8,9):
        for _column in range(2,9):
            worksheet.cell(row=_row, column=_column).border = under_border
    
    for _row in range(54,55):
        for _column in range(2,9):
            worksheet.cell(row=_row, column=_column).border = under_border

    for _row in range(58,59):
        for _column in range(2,9):
            worksheet.cell(row=_row, column=_column).border = under_border

    for _row in range(12,14):
        for _column in range(7,9):
            worksheet.cell(row=_row, column=_column).border = under_border

    
    for _column in range(2,9):
        worksheet.cell(row=11, column=_column).border = under_border

    #ALIGNMENTS
    horizontal_center = Alignment(horizontal='center')
    wrap_center = Alignment(wrap_text='center')
    horizontal_right = Alignment(horizontal='right')
 

    for _row in range(1,11):
        worksheet.cell(row=_row, column=2).alignment = horizontal_center
    
    for _row in range(15,49):
        for _column in range(2,9):
            worksheet.cell(row=_row, column=_column).alignment = horizontal_center
     
    for _row in range(12,14):
        worksheet.cell(row=_row, column=2).alignment = wrap_center

    for _row in range(50,53):
        worksheet.cell(row=_row, column=2).alignment = wrap_center
    
    for _row in range(54,60):
        worksheet.cell(row=_row, column=2).alignment = horizontal_center

    worksheet.cell(row=48, column=2).alignment = horizontal_right
    worksheet.cell(row=15, column=3).alignment = horizontal_center
    worksheet.cell(row=15, column=5).alignment = horizontal_center
    
    #FONTS
    name_font = Font(name='Bodoni MT Black', b=True)
    date_font = Font(size=14, b=True)
    font_18_bold = Font(size=18, b=True)
    font_14_bold = Font(size=14, b=True)
    font_9 = Font(size=9)
    
    worksheet.cell(row=8, column=2).font = name_font
    worksheet.cell(row=2, column=2).font = font_18_bold
    worksheet.cell(row=3, column=2).font = font_9
    worksheet.cell(row=4, column=2).font = font_9
    worksheet.cell(row=6, column=2).font = font_14_bold
    worksheet.cell(row=11, column=2).font = font_9
    worksheet.cell(row=12, column=2).font = font_9
    worksheet.cell(row=13, column=2).font = font_9
    worksheet.cell(row=11, column=6).font = date_font

    #COLUMN WIDTH
    worksheet.column_dimensions['A'].width = 1
    worksheet.column_dimensions['I'].width = 1
    worksheet.column_dimensions['B'].width = 8
    worksheet.column_dimensions['C'].width = 8
    worksheet.column_dimensions['D'].width = 8
    worksheet.column_dimensions['E'].width = 8
    worksheet.column_dimensions['F'].width = 8
    worksheet.column_dimensions['G'].width = 8
    worksheet.column_dimensions['H'].width = 9
    




def log_user_to_workbook(in_or_out, worker, clock_in_pos, midnight):
    today = datetime.now()
    
    month = today.strftime("%B")
    workbook = load_workbook(f"dtr/output/{month}_{today.year}.xlsx")
    time = today.strftime("%I:%M")
    if worker in workbook.sheetnames:
        sheet = workbook[worker]
        header_started = sheet[clock_in_pos]
        if not midnight:
            header_started.value = f"{time}"
        else:
            header_started.value = f"{time} MN"


        workbook.save(f"dtr/output/{month}_{today.year}.xlsx")
        print(f'{worker} logged {in_or_out} @ {time} today.')
    else:           
        workbook.create_sheet(worker)
        print(f'{worker} added to workbook daily time record.')
        sheet = workbook[worker]
        create_dtr(sheet, worker)

        header_started = sheet[clock_in_pos]
        
        if not midnight:
            header_started.value = f"{time}"
        else:
            header_started.value = f"{time} MN"

        print(f'{worker} logged {in_or_out} @ {time} today.')

        workbook.save(f"dtr/output/{month}_{today.year}.xlsx")
    




# :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
# :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
# :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
# :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
def worklogger(worker):
    
    today = datetime.now()
    month = today.strftime("%B")

    in_or_out = "IN"

    #IF NO WORKBOOK CREATED
    if not path.exists(f"dtr/output/{month}_{today.year}.xlsx"):
        create_new_workbook()
        
        if today.strftime("%p") == 'AM':
            clock_in_pos = f'C{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

        elif today.strftime("%p") == 'PM':
            clock_in_pos = f'E{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

        elif today.strftime("%p") == 'AM' and today.strftime("%I") == '12':
            clock_in_pos = f'E{16 + today.day}'
            midnight = True
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

    #IF WORKBOOK EXIST
    else:
        if today.strftime("%p") == 'AM':
            clock_in_pos = f'C{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)
            

        elif today.strftime("%p") == 'PM':
            clock_in_pos = f'E{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)
        
        elif today.strftime("%p") == 'AM' and today.strftime("%I") == '12':
            clock_in_pos = f'E{16 + today.day}'
            midnight = True
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

    

def worklogger_out(worker):
    today = datetime.now()
    month = today.strftime("%B")

    clock_in_pos = f'B{today.day}'
    in_or_out = "OUT"
    #IF NO WORKBOOK CREATED
    if not path.exists(f"dtr/output/{month}_{today.year}.xlsx"):
        create_new_workbook()
        
        if today.strftime("%p") == 'AM':
            clock_in_pos = f'D{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

        elif today.strftime("%p") == 'PM':
            clock_in_pos = f'F{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)
        
        elif today.strftime("%p") == 'AM' and today.strftime("%I") == '12':
            clock_in_pos = f'F{16 + today.day}'
            midnight = True
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

    #IF WORKBOOK EXIST
    else:
        if today.strftime("%p") == 'AM':
            clock_in_pos = f'D{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

        elif today.strftime("%p") == 'PM':
            clock_in_pos = f'F{16 + today.day}'
            midnight = False
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)
        
        elif today.strftime("%p") == 'AM' and today.strftime("%I") == '12':
            clock_in_pos = f'F{16 + today.day}'
            midnight = True
            log_user_to_workbook(in_or_out, worker,clock_in_pos, midnight)

