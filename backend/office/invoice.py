import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


from datetime import datetime
from .models import Invoice, AdVersion, Advertiser, InvoiceTransmittal
from station.models import User
from django.core.exceptions import ObjectDoesNotExist





def create_invoice(invoice_no):

    invoice = Invoice.objects.get(invoice_no=invoice_no)
        
   

    template = os.path.abspath('office/invoice/template-invoice.xlsx')

    workbook = load_workbook(template)

    sheet = workbook['subject']

    
    
    #TITLE
    company = invoice.advertiser.name.title()

    #ADDRESS
    address = invoice.advertiser.address

    #DATE
    date = invoice.invoice_date.strftime('%d/%m/%Y')

    #PARTICULAR
    app_date_from = invoice.applicable_month_from
    app_date_to = invoice.applicable_month_to

    if app_date_from.strftime('%Y') == app_date_to.strftime('%Y'):
        particular = app_date_from.strftime('%B %d') + '-' + app_date_to.strftime('%B %d, %Y')
    else:
        particular = app_date_from.strftime('%B %d, %Y') + '-' + app_date_to.strftime('%B %d, %Y')

    #AMOUNT
    amount = invoice.amount

    #IN-CHARGE
    #
    in_charge = 'Florence Maynopas-Dumalag'

    make_invoice(sheet, company, address, date, particular, amount, in_charge)

    workbook.save(f'office/invoice/generated/{company} inv#{invoice_no}.xlsx')

    return f'office/invoice/generated/{company} inv#{invoice_no}.xlsx'





def make_invoice(sheet, advertiser, address, date, particular, amount, in_charge):
    #TITLE
    sheet['A4'] = advertiser

    #ADDRESS
    sheet['A5'] = address

    #DATE
    sheet['I4'] = date

    #PARTICULAR
    sheet['A11'] = particular

    #AMOUNT
    sheet['F12'] = amount
    sheet['I12'] = amount

    #IN-CHARGE
    sheet['A24'] = in_charge


def invoice_append(sheet, monthyear, invoice_date, invoice_no, advertiser, particular, amount):
    print(sheet.max_row)
    #MONTHYEAR
    sheet["B4"] = monthyear

    sheet.append([invoice_date, invoice_no, advertiser, particular, amount])
    sheet[f"A{sheet.max_row}"].alignment = Alignment(horizontal='left')
    sheet[f"B{sheet.max_row}"].alignment = Alignment(horizontal='center')
    sheet[f"E{sheet.max_row}"].number_format = '₱#,##0.00'
    sheet[f"E{sheet.max_row}"].alignment = Alignment(horizontal='left')

def collection_append_add(sheet,monthyear,invoice_date,invoice_no,amount,advertiser,ae,or_date,or_number,applicable_month,remarks):
    sheet["B3"] = monthyear

    sheet.append([invoice_date,invoice_no,amount,advertiser,ae,or_date,or_number,applicable_month,remarks])
    sheet[f"F{sheet.max_row}"].alignment = Alignment(horizontal='center')
    sheet[f"G{sheet.max_row}"].alignment = Alignment(horizontal='center')
    sheet[f"I{sheet.max_row}"].alignment = Alignment(horizontal='center')